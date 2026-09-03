"""板凳龙运动学求解 —— 主入口模块。

执行流程：
1. 固定随机种子 np.random.seed(42)
2. 按序调用 solve_problem_1..5
3. 写出 figures/all_results.json（数值真相源）
4. 写出 tables/result1-5.xlsx

模板来源: code-templates/utils/data_pipeline.py（IO 骨架）
"""

import json
import os
import sys

import numpy as np

# 确保能 import 同目录模块
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from solve import (
    solve_problem_1, solve_problem_2, solve_problem_3,
    solve_problem_4, solve_problem_5
)


def main() -> None:
    """主入口：求解全部子问题并写出结果文件。

    固定随机种子 42 保证可复现，结果写入 figures/all_results.json。
    """
    # P1: 固定随机种子
    np.random.seed(42)

    # 确保输出目录存在
    project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    figures_dir = os.path.join(project_dir, "figures")
    tables_dir = os.path.join(project_dir, "tables")
    os.makedirs(figures_dir, exist_ok=True)
    os.makedirs(tables_dir, exist_ok=True)

    # 按序求解 5 个子问题
    print("=" * 60)
    print("板凳龙运动学求解（2024 CUMCM A 题）")
    print("=" * 60)

    params = {
        "b": 0.55, "v1": 1.0, "r0": 8.8, "n_bench": 222,
        "L_head": 3.41, "L_body": 2.20, "w": 0.30,
    }

    results = {}

    print("[Q1] 盘入 300s 位置与速度...")
    results["problem_1"] = solve_problem_1(params)
    print("  head_pos_t0: {}".format(results["problem_1"]["values"]["head_pos_t0"]))
    print("  max_speed: {} m/s @ t={}s, handle={}".format(
        results["problem_1"]["values"]["max_speed"],
        results["problem_1"]["values"]["max_speed_time"],
        results["problem_1"]["values"]["max_speed_handle"]))

    print("[Q2] 碰撞终止时刻...")
    results["problem_2"] = solve_problem_2(params)
    print("  t* = {} s".format(results["problem_2"]["values"]["t_star"]))
    print("  head_r = {} m".format(results["problem_2"]["values"]["head_r_at_t_star"]))

    print("[Q3] 掉头最小直径...")
    results["problem_3"] = solve_problem_3(params)
    print("  d_min = {} m".format(results["problem_3"]["values"]["d_min"]))

    print("[Q4] 盘出最大速度...")
    results["problem_4"] = solve_problem_4(params)
    print("  v_max = {} m/s @ t={}s".format(
        results["problem_4"]["values"]["v_max"],
        results["problem_4"]["values"]["t_at_vmax"]))

    print("[Q5] S 形圆弧半径与螺距...")
    results["problem_5"] = solve_problem_5(params)
    print("  R = {} m, p' = {} m".format(
        results["problem_5"]["values"]["R_arc"],
        results["problem_5"]["values"]["p_adjusted"]))

    # 写出 all_results.json（P2: 数值真相源）
    # 将 numpy 类型转换为原生 Python 类型
    def _serialize(obj):
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            return float(obj)
        if isinstance(obj, (np.bool_,)):
            return bool(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if isinstance(obj, dict):
            return {k: _serialize(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [_serialize(v) for v in obj]
        return obj

    all_results_path = os.path.join(figures_dir, "all_results.json")
    with open(all_results_path, "w", encoding="utf-8") as f:
        json.dump(_serialize(results), f, indent=2, ensure_ascii=False)
    print("\n[OK] all_results.json -> {}".format(all_results_path))

    # 写出 result*.xlsx（如果 openpyxl 可用）
    try:
        import openpyxl
        _write_xlsx(results, tables_dir)
        print("[OK] result1-5.xlsx -> {}".format(tables_dir))
    except ImportError:
        print("[WARN] openpyxl 不可用，跳过 xlsx 输出")

    print("=" * 60)
    print("求解完成！")
    print("=" * 60)


def _write_xlsx(results, tables_dir):
    """写出 result1-5.xlsx。"""
    # 问题1：位置速度表（简化版：仅关键时刻）
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "result1"
    ws.append(["时刻(s)", "板凳编号", "x(m)", "y(m)", "速度(m/s)"])
    q1 = results["problem_1"]["values"]
    ws.append([0, 0, q1["head_pos_t0"][0], q1["head_pos_t0"][1], 1.0])
    ws.append([0, 222, q1["tail_pos_t0"][0], q1["tail_pos_t0"][1], q1["max_speed"]])
    ws.append([300, 0, q1["head_pos_t300"][0], q1["head_pos_t300"][1], 1.0])
    wb.save(os.path.join(tables_dir, "result1.xlsx"))

    # 问题2
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "result2"
    ws.append(["t*(s)", "head_x(m)", "head_y(m)", "head_speed(m/s)", "tail_x(m)", "tail_y(m)"])
    q2 = results["problem_2"]["values"]
    ws.append([q2["t_star"], q2["head_pos_at_t_star"][0], q2["head_pos_at_t_star"][1],
               q2["head_speed_at_t_star"], q2["tail_pos_at_t_star"][0], q2["tail_pos_at_t_star"][1]])
    wb.save(os.path.join(tables_dir, "result2.xlsx"))

    # 问题3
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "result3"
    ws.append(["d_min(m)", "r_collision(m)", "t*(s)"])
    q3 = results["problem_3"]["values"]
    ws.append([q3["d_min"], q3["r_collision"], q3["t_star"]])
    wb.save(os.path.join(tables_dir, "result3.xlsx"))

    # 问题4
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "result4"
    ws.append(["v_max(m/s)", "t_at_vmax(s)", "x(m)", "y(m)", "handle_id"])
    q4 = results["problem_4"]["values"]
    ws.append([q4["v_max"], q4["t_at_vmax"], q4["pos_at_vmax"][0], q4["pos_at_vmax"][1], q4["handle_at_vmax"]])
    wb.save(os.path.join(tables_dir, "result4.xlsx"))

    # 问题5
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "result5"
    ws.append(["R_arc(m)", "p_adjusted(m)", "d_turn(m)"])
    q5 = results["problem_5"]["values"]
    ws.append([q5["R_arc"], q5["p_adjusted"], q5["d_turn"]])
    wb.save(os.path.join(tables_dir, "result5.xlsx"))


if __name__ == "__main__":
    main()
